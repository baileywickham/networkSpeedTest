import speedtest
from openpyxl import Workbook
import datetime

servers = []

def main():

        # s = speedtest.Speedtest()
        # s.get_servers(servers)
        # s.get_best_server()
        # s.download()
        # s.upload()
        # results = s.results.dict()

        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted

        ws['A2'] = datetime.datetime.now()

        # Save the file
        wb.save("speed.xlsx")
        for row in ws.rows:
            print(row)

if __name__ == "__main__":
    main()